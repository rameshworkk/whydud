"""Generate ERD JSON and Data Dictionary Excel from the Whydud database schema."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# ERD JSON
# =============================================================================

erd = {
    "project": "Whydud Platform",
    "description": "India-first product intelligence and trust platform — full database ERD",
    "database": {
        "engine": "PostgreSQL 16 + TimescaleDB",
        "extensions": ["timescaledb", "uuid-ossp", "pgcrypto"],
        "schemas": ["public", "users", "email_intel", "scoring", "tco", "community", "admin"]
    },
    "schemas": {
        "public": {
            "description": "Products, marketplaces, listings, reviews, deals, pricing, search, scraping",
            "tables": ["Marketplace", "Category", "Brand", "Product", "Seller", "ProductListing",
                        "BankCard", "CompareSession", "RecentlyViewed", "StockAlert",
                        "CategoryPreferenceSchema", "PriceSnapshot", "MarketplaceOffer",
                        "PriceAlert", "ClickEvent", "BackfillProduct", "Review", "ReviewVote",
                        "ReviewerProfile", "Deal", "SearchLog", "ScraperJob"]
        },
        "users": {
            "description": "Accounts, auth, payments, notifications, wishlists, rewards, TCO profiles",
            "tables": ["User", "WhydudEmail", "OAuthConnection", "PaymentMethod",
                        "ReservedUsername", "Notification", "NotificationPreference",
                        "MarketplacePreference", "PurchasePreference", "Wishlist",
                        "WishlistItem", "RewardPointsLedger", "RewardBalance",
                        "GiftCardCatalog", "GiftCardRedemption", "UserTCOProfile"]
        },
        "email_intel": {
            "description": "Inbox emails, parsed orders, refunds, return windows, subscriptions",
            "tables": ["InboxEmail", "EmailSource", "ParsedOrder", "RefundTracking",
                        "ReturnWindow", "DetectedSubscription"]
        },
        "scoring": {
            "description": "DudScore config, history, brand trust",
            "tables": ["DudScoreConfig", "DudScoreHistory", "BrandTrustScore"]
        },
        "tco": {
            "description": "Total Cost of Ownership models and city reference data",
            "tables": ["TCOModel", "CityReferenceData"]
        },
        "community": {
            "description": "Discussion threads, replies, votes",
            "tables": ["DiscussionThread", "DiscussionReply", "DiscussionVote"]
        },
        "admin": {
            "description": "Audit logs, moderation queue, scraper runs, site config",
            "tables": ["AuditLog", "ModerationQueue", "ScraperRun", "SiteConfig"]
        }
    },
    "entities": {
        "User": {
            "schema": "users",
            "django_app": "accounts",
            "description": "Primary user identity. Custom auth model (email-based, no username). Supports role hierarchy and subscription tiers.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": "Unique user identifier"},
                {"name": "email", "type": "EmailField", "unique": True, "nullable": False, "description": "Login email address (primary auth credential)"},
                {"name": "name", "type": "CharField(150)", "nullable": False, "description": "User display name"},
                {"name": "avatar_url", "type": "URLField", "nullable": True, "description": "Profile picture URL"},
                {"name": "role", "type": "CharField(20)", "nullable": False, "default": "registered", "choices": ["registered","connected","premium","moderator","senior_moderator","data_ops","fraud_analyst","trust_engineer","admin","super_admin"], "description": "Access role"},
                {"name": "subscription_tier", "type": "CharField(10)", "nullable": False, "default": "free", "choices": ["free","premium"], "description": "Subscription plan"},
                {"name": "has_whydud_email", "type": "BooleanField", "nullable": False, "default": False, "description": "Whether user has activated a @whyd.* shopping email"},
                {"name": "referral_code", "type": "CharField(12)", "unique": True, "nullable": False, "description": "Auto-generated invite code for referral program"},
                {"name": "referred_by", "type": "FK(User)", "nullable": True, "fk_to": "User", "description": "User who referred this account (self-referential)"},
                {"name": "trust_score", "type": "DecimalField(5,2)", "nullable": True, "description": "Platform-computed trust score (0.00-100.00)"},
                {"name": "is_suspended", "type": "BooleanField", "nullable": False, "default": False, "description": "Account suspension flag"},
                {"name": "deletion_requested_at", "type": "DateTimeField", "nullable": True, "description": "Timestamp of GDPR/account deletion request"},
                {"name": "is_active", "type": "BooleanField", "nullable": False, "default": True, "description": "Django auth: can log in"},
                {"name": "is_staff", "type": "BooleanField", "nullable": False, "default": False, "description": "Django auth: can access admin site"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": "Account creation timestamp"},
                {"name": "updated_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": "Last profile update timestamp"}
            ],
            "indexes": ["email (unique)", "referral_code (unique)", "role", "subscription_tier", "created_at"],
            "unique_constraints": ["email", "referral_code"]
        },
        "WhydudEmail": {
            "schema": "users",
            "django_app": "accounts",
            "description": "Shopping email addresses under @whyd.in / @whyd.click / @whyd.shop. Users register these on marketplaces to enable purchase tracking.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "on_delete": "CASCADE", "description": "Owner of this shopping email"},
                {"name": "username", "type": "CharField(64)", "nullable": False, "description": "Local part of the email (e.g. ravi in ravi@whyd.in)"},
                {"name": "domain", "type": "CharField(20)", "nullable": False, "choices": ["whyd.in","whyd.click","whyd.shop"], "description": "Domain suffix"},
                {"name": "is_active", "type": "BooleanField", "nullable": False, "default": True, "description": "Whether this email is actively receiving mail"},
                {"name": "total_emails_received", "type": "IntegerField", "nullable": False, "default": 0, "description": "Running counter of inbound emails"},
                {"name": "total_orders_detected", "type": "IntegerField", "nullable": False, "default": 0, "description": "Running counter of parsed purchase orders"},
                {"name": "last_email_received_at", "type": "DateTimeField", "nullable": True, "description": "Timestamp of most recent email"},
                {"name": "onboarding_complete", "type": "BooleanField", "nullable": False, "default": False, "description": "Whether user has completed marketplace registration guide"},
                {"name": "marketplaces_registered", "type": "JSONField", "nullable": False, "default": [], "description": "List of marketplace slugs where this email is registered"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ],
            "unique_constraints": [["username", "domain"]]
        },
        "OAuthConnection": {
            "schema": "users",
            "django_app": "accounts",
            "description": "OAuth2 tokens for external services (currently Google/Gmail). Tokens encrypted at rest with AES-256-GCM.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "on_delete": "CASCADE", "description": "Token owner"},
                {"name": "provider", "type": "CharField(20)", "nullable": False, "choices": ["google"], "description": "OAuth provider"},
                {"name": "provider_user_id", "type": "CharField(128)", "nullable": False, "description": "Provider's unique user ID"},
                {"name": "access_token_encrypted", "type": "BinaryField", "nullable": False, "encrypted": "AES-256-GCM", "description": "Encrypted OAuth access token"},
                {"name": "refresh_token_encrypted", "type": "BinaryField", "nullable": False, "encrypted": "AES-256-GCM", "description": "Encrypted OAuth refresh token"},
                {"name": "token_expires_at", "type": "DateTimeField", "nullable": False, "description": "Access token expiry"},
                {"name": "scopes", "type": "JSONField", "nullable": False, "default": [], "description": "Granted OAuth scopes"},
                {"name": "connected_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": "When OAuth was first connected"},
                {"name": "last_sync_at", "type": "DateTimeField", "nullable": True, "description": "Last successful data sync"},
                {"name": "status", "type": "CharField(10)", "nullable": False, "default": "active", "choices": ["active","expired","revoked"], "description": "Connection health"}
            ],
            "unique_constraints": [["provider", "provider_user_id"]]
        },
        "PaymentMethod": {
            "schema": "users",
            "django_app": "accounts",
            "description": "Card vault storing bank/card metadata for payment optimization. NO card numbers, CVV, or expiry stored.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "on_delete": "CASCADE", "description": "Card owner"},
                {"name": "method_type", "type": "CharField(20)", "nullable": False, "choices": ["credit_card","debit_card","upi","wallet","membership"], "description": "Payment type"},
                {"name": "bank_name", "type": "CharField(100)", "nullable": True, "description": "Issuing bank name"},
                {"name": "card_variant", "type": "CharField(100)", "nullable": True, "description": "Card product name (e.g. Regalia, Millennia)"},
                {"name": "card_network", "type": "CharField(20)", "nullable": True, "choices": ["visa","mastercard","rupay","amex","diners"], "description": "Card network"},
                {"name": "wallet_provider", "type": "CharField(50)", "nullable": True, "description": "E-wallet name"},
                {"name": "wallet_balance", "type": "DecimalField(12,2)", "nullable": True, "description": "Self-reported wallet balance (paisa)"},
                {"name": "upi_app", "type": "CharField(50)", "nullable": True, "description": "UPI app name"},
                {"name": "upi_bank", "type": "CharField(100)", "nullable": True, "description": "UPI-linked bank"},
                {"name": "membership_type", "type": "CharField(50)", "nullable": True, "description": "Platform membership (e.g. Amazon Prime)"},
                {"name": "emi_eligible", "type": "BooleanField", "nullable": False, "default": False, "description": "Whether this card supports EMI"},
                {"name": "nickname", "type": "CharField(50)", "nullable": True, "description": "User-assigned label"},
                {"name": "is_preferred", "type": "BooleanField", "nullable": False, "default": False, "description": "User's primary payment method"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ],
            "security_note": "NEVER stores card number, CVV, or expiry. Metadata only."
        },
        "ReservedUsername": {
            "schema": "users",
            "django_app": "accounts",
            "description": "Seeded list of usernames blocked from registration as shopping emails.",
            "primary_key": {"field": "id", "type": "int"},
            "fields": [
                {"name": "id", "type": "IntegerField", "pk": True, "nullable": False, "description": "Auto-increment"},
                {"name": "username", "type": "CharField(64)", "unique": True, "nullable": False, "description": "Reserved local part"}
            ]
        },
        "Notification": {
            "schema": "users",
            "django_app": "accounts",
            "description": "In-app and email notification records.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "on_delete": "CASCADE", "description": "Recipient"},
                {"name": "type", "type": "CharField(30)", "nullable": False, "choices": ["price_drop","return_window","refund_delay","back_in_stock","review_upvote","price_alert","discussion_reply","level_up","points_earned","subscription_renewal","order_detected"], "description": "Notification category"},
                {"name": "title", "type": "CharField(200)", "nullable": False, "description": "Notification headline"},
                {"name": "body", "type": "TextField", "nullable": False, "description": "Full notification text"},
                {"name": "action_url", "type": "URLField", "nullable": True, "description": "Deep link to relevant page"},
                {"name": "action_label", "type": "CharField(50)", "nullable": True, "description": "CTA button text"},
                {"name": "entity_type", "type": "CharField(50)", "nullable": True, "description": "Polymorphic reference type"},
                {"name": "entity_id", "type": "CharField(50)", "nullable": True, "description": "Polymorphic reference ID"},
                {"name": "metadata", "type": "JSONField", "nullable": False, "default": {}, "description": "Additional context data"},
                {"name": "is_read", "type": "BooleanField", "nullable": False, "default": False, "description": "Read status"},
                {"name": "email_sent", "type": "BooleanField", "nullable": False, "default": False, "description": "Whether email was dispatched"},
                {"name": "email_sent_at", "type": "DateTimeField", "nullable": True, "description": "Email dispatch timestamp"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ]
        },
        "NotificationPreference": {
            "schema": "users",
            "django_app": "accounts",
            "description": "Per-user channel toggles for each notification type.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "on_delete": "CASCADE", "description": ""},
                {"name": "notification_type", "type": "CharField(30)", "nullable": False, "description": "Matches Notification.type values"},
                {"name": "in_app_enabled", "type": "BooleanField", "nullable": False, "default": True, "description": "Show in-app notifications"},
                {"name": "email_enabled", "type": "BooleanField", "nullable": False, "default": True, "description": "Send email notifications"}
            ],
            "unique_constraints": [["user", "notification_type"]]
        },
        "MarketplacePreference": {
            "schema": "users",
            "django_app": "accounts",
            "description": "User's preferred marketplaces for default filtering.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "on_delete": "CASCADE", "description": ""},
                {"name": "marketplace_slug", "type": "CharField(30)", "nullable": False, "description": "References Marketplace.slug"},
                {"name": "is_enabled", "type": "BooleanField", "nullable": False, "default": True, "description": "Include in default filters"}
            ]
        },
        "PurchasePreference": {
            "schema": "users",
            "django_app": "accounts",
            "description": "Per-category buying preferences for personalized recommendations.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "on_delete": "CASCADE", "description": ""},
                {"name": "category", "type": "FK(Category)", "nullable": False, "fk_to": "Category", "description": "Product category"},
                {"name": "answers", "type": "JSONField", "nullable": False, "description": "Questionnaire responses (schema from CategoryPreferenceSchema)"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ]
        },
        "Marketplace": {
            "schema": "public",
            "django_app": "products",
            "description": "Supported Indian e-commerce marketplaces.",
            "primary_key": {"field": "id", "type": "int"},
            "fields": [
                {"name": "id", "type": "IntegerField", "pk": True, "nullable": False, "description": "Auto-increment"},
                {"name": "slug", "type": "SlugField", "unique": True, "nullable": False, "description": "URL-safe identifier (e.g. amazon-in, flipkart)"},
                {"name": "name", "type": "CharField(100)", "nullable": False, "description": "Display name"},
                {"name": "base_url", "type": "URLField", "nullable": False, "description": "Marketplace homepage URL"},
                {"name": "affiliate_tag", "type": "CharField(50)", "nullable": True, "description": "Affiliate tracking ID"},
                {"name": "affiliate_param", "type": "CharField(30)", "nullable": True, "description": "URL parameter name for affiliate (e.g. tag for Amazon)"},
                {"name": "scraper_status", "type": "CharField(10)", "nullable": False, "default": "active", "choices": ["active","paused","broken"], "description": "Scraper health"}
            ]
        },
        "Category": {
            "schema": "public",
            "django_app": "products",
            "description": "Hierarchical product taxonomy. Self-referential tree with optional TCO model.",
            "primary_key": {"field": "id", "type": "int"},
            "fields": [
                {"name": "id", "type": "IntegerField", "pk": True, "nullable": False, "description": "Auto-increment"},
                {"name": "slug", "type": "SlugField", "unique": True, "nullable": False, "description": "URL-safe identifier"},
                {"name": "name", "type": "CharField(100)", "nullable": False, "description": "Display name"},
                {"name": "parent", "type": "FK(Category)", "nullable": True, "fk_to": "Category", "description": "Parent category for tree hierarchy (self-ref)"},
                {"name": "spec_schema", "type": "JSONField", "nullable": False, "default": {}, "description": "JSON Schema defining expected product specs"},
                {"name": "level", "type": "IntegerField", "nullable": False, "default": 0, "description": "Depth in category tree (0 = root)"},
                {"name": "has_tco_model", "type": "BooleanField", "nullable": False, "default": False, "description": "Whether TCO calculator exists for this category"},
                {"name": "product_count", "type": "IntegerField", "nullable": False, "default": 0, "description": "Denormalized count of active products"}
            ]
        },
        "Brand": {
            "schema": "public",
            "django_app": "products",
            "description": "Product brand registry. Supports aliases for fuzzy matching across marketplaces.",
            "primary_key": {"field": "id", "type": "int"},
            "fields": [
                {"name": "id", "type": "IntegerField", "pk": True, "nullable": False, "description": "Auto-increment"},
                {"name": "slug", "type": "SlugField", "unique": True, "nullable": False, "description": "URL-safe identifier"},
                {"name": "name", "type": "CharField(100)", "nullable": False, "description": "Canonical brand name"},
                {"name": "aliases", "type": "JSONField", "nullable": False, "default": [], "description": "Alternative names/spellings for cross-marketplace matching"},
                {"name": "verified", "type": "BooleanField", "nullable": False, "default": False, "description": "Whether brand identity has been manually verified"},
                {"name": "logo_url", "type": "URLField", "nullable": True, "description": "Brand logo image URL"}
            ]
        },
        "Product": {
            "schema": "public",
            "django_app": "products",
            "description": "Canonical product entity aggregated across all marketplace listings. Central node in the data model.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "slug", "type": "SlugField", "unique": True, "nullable": False, "description": "SEO-friendly URL slug"},
                {"name": "title", "type": "CharField(500)", "nullable": False, "description": "Canonical product title"},
                {"name": "brand", "type": "FK(Brand)", "nullable": True, "fk_to": "Brand", "description": "Product manufacturer"},
                {"name": "category", "type": "FK(Category)", "nullable": True, "fk_to": "Category", "description": "Product taxonomy classification"},
                {"name": "description", "type": "TextField", "nullable": True, "description": "Product description (may be scraped)"},
                {"name": "specs", "type": "JSONField", "nullable": False, "default": {}, "description": "Structured specifications (keys from Category.spec_schema)"},
                {"name": "images", "type": "JSONField", "nullable": False, "default": [], "description": "Array of image URLs (first = primary)"},
                {"name": "dud_score", "type": "DecimalField(5,2)", "nullable": True, "description": "Whydud proprietary trust/value score (0.00-100.00)"},
                {"name": "dud_score_confidence", "type": "DecimalField(3,2)", "nullable": True, "description": "Score confidence (0.00-1.00)"},
                {"name": "dud_score_updated_at", "type": "DateTimeField", "nullable": True, "description": "Last DudScore computation timestamp"},
                {"name": "avg_rating", "type": "DecimalField(3,2)", "nullable": True, "description": "Weighted average rating across all listings"},
                {"name": "total_reviews", "type": "IntegerField", "nullable": False, "default": 0, "description": "Total review count across all listings"},
                {"name": "lowest_price_ever", "type": "DecimalField(12,2)", "nullable": True, "description": "Historical lowest price (paisa)"},
                {"name": "lowest_price_date", "type": "DateField", "nullable": True, "description": "Date of lowest price"},
                {"name": "current_best_price", "type": "DecimalField(12,2)", "nullable": True, "description": "Current lowest available price (paisa)"},
                {"name": "current_best_marketplace", "type": "CharField(30)", "nullable": True, "description": "Marketplace slug with current best price"},
                {"name": "is_refurbished", "type": "BooleanField", "nullable": False, "default": False, "description": "Whether this is refurbished/renewed"},
                {"name": "status", "type": "CharField(15)", "nullable": False, "default": "active", "choices": ["active","discontinued","pending"], "description": "Lifecycle status"},
                {"name": "first_seen_at", "type": "DateTimeField", "nullable": True, "description": "When first scraped/detected"},
                {"name": "last_scraped_at", "type": "DateTimeField", "nullable": True, "description": "Most recent scrape"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": "DB creation time"},
                {"name": "updated_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": "Last modification time"}
            ]
        },
        "Seller": {
            "schema": "public",
            "django_app": "products",
            "description": "Marketplace seller profiles.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "marketplace", "type": "FK(Marketplace)", "nullable": False, "fk_to": "Marketplace", "description": "Which marketplace this seller operates on"},
                {"name": "external_seller_id", "type": "CharField(200)", "nullable": False, "description": "Marketplace's seller identifier"},
                {"name": "name", "type": "CharField(300)", "nullable": False, "description": "Seller display name"},
                {"name": "avg_rating", "type": "DecimalField(3,2)", "nullable": True, "description": "Seller's average rating"},
                {"name": "total_ratings", "type": "IntegerField", "nullable": False, "default": 0, "description": "Total number of seller ratings"},
                {"name": "positive_pct", "type": "DecimalField(5,2)", "nullable": True, "description": "Percentage of positive ratings"},
                {"name": "ships_from", "type": "CharField(200)", "nullable": True, "description": "Shipping origin location"},
                {"name": "fulfilled_by", "type": "CharField(200)", "nullable": True, "description": "Fulfillment entity"},
                {"name": "is_verified", "type": "BooleanField", "nullable": False, "default": False, "description": "Seller identity verified"},
                {"name": "seller_since", "type": "DateField", "nullable": True, "description": "Date seller started on marketplace"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ],
            "unique_constraints": [["marketplace", "external_seller_id"]]
        },
        "ProductListing": {
            "schema": "public",
            "django_app": "products",
            "description": "A specific product listed on a specific marketplace by a specific seller.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "product", "type": "FK(Product)", "nullable": True, "fk_to": "Product", "description": "Canonical product (null if unmatched)"},
                {"name": "marketplace", "type": "FK(Marketplace)", "nullable": False, "fk_to": "Marketplace", "description": "Source marketplace"},
                {"name": "seller", "type": "FK(Seller)", "nullable": True, "fk_to": "Seller", "description": "Selling merchant"},
                {"name": "external_id", "type": "CharField(100)", "nullable": False, "description": "Marketplace's product ID (e.g. Amazon ASIN)"},
                {"name": "external_url", "type": "URLField(2000)", "nullable": False, "description": "Direct URL on marketplace"},
                {"name": "affiliate_url", "type": "URLField(2000)", "nullable": True, "description": "URL with affiliate tracking"},
                {"name": "title", "type": "CharField(500)", "nullable": False, "description": "Listing title as shown on marketplace"},
                {"name": "current_price", "type": "DecimalField(12,2)", "nullable": True, "description": "Current selling price (paisa)"},
                {"name": "mrp", "type": "DecimalField(12,2)", "nullable": True, "description": "Maximum Retail Price (paisa)"},
                {"name": "discount_pct", "type": "DecimalField(5,2)", "nullable": True, "description": "Discount percentage"},
                {"name": "in_stock", "type": "BooleanField", "nullable": False, "default": True, "description": "Stock availability"},
                {"name": "rating", "type": "DecimalField(3,2)", "nullable": True, "description": "Listing-specific rating"},
                {"name": "review_count", "type": "IntegerField", "nullable": False, "default": 0, "description": "Listing-specific review count"},
                {"name": "match_confidence", "type": "DecimalField(3,2)", "nullable": True, "description": "Confidence of product matching (0.00-1.00)"},
                {"name": "match_method", "type": "CharField(30)", "nullable": True, "description": "How matching was done (e.g. ean, title_similarity)"},
                {"name": "last_scraped_at", "type": "DateTimeField", "nullable": True, "description": "Last successful scrape"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ],
            "unique_constraints": [["marketplace", "external_id"]]
        },
        "BankCard": {
            "schema": "public",
            "django_app": "products",
            "description": "Reference data for Indian bank debit/credit cards. Used to match offers to user payment methods.",
            "primary_key": {"field": "id", "type": "int"},
            "fields": [
                {"name": "id", "type": "IntegerField", "pk": True, "nullable": False, "description": ""},
                {"name": "bank_slug", "type": "SlugField", "nullable": False, "description": "URL-safe bank identifier"},
                {"name": "bank_name", "type": "CharField(100)", "nullable": False, "description": "Display name"},
                {"name": "card_variant", "type": "CharField(100)", "nullable": False, "description": "Card product name"},
                {"name": "card_type", "type": "CharField(10)", "nullable": False, "choices": ["credit","debit"], "description": "Card type"},
                {"name": "card_network", "type": "CharField(20)", "nullable": False, "description": "Network (Visa, Mastercard, RuPay, Amex)"},
                {"name": "is_co_branded", "type": "BooleanField", "nullable": False, "default": False, "description": "Co-branded card flag"},
                {"name": "co_brand_partner", "type": "CharField(100)", "nullable": True, "description": "Co-brand partner name"},
                {"name": "default_cashback_pct", "type": "DecimalField(5,2)", "nullable": False, "default": 0, "description": "Default cashback rate"},
                {"name": "logo_url", "type": "URLField", "nullable": True, "description": "Card logo image"}
            ],
            "unique_constraints": [["bank_slug", "card_variant", "card_type"]]
        },
        "CompareSession": {
            "schema": "public",
            "django_app": "products",
            "description": "Tracks products a user is actively comparing.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": True, "fk_to": "User", "description": "Logged-in user (null for anonymous)"},
                {"name": "session_id", "type": "CharField(64)", "nullable": True, "description": "Browser session ID for anonymous users"},
                {"name": "product_ids", "type": "JSONField", "nullable": False, "default": [], "description": "Array of product UUIDs being compared"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""},
                {"name": "updated_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ]
        },
        "RecentlyViewed": {
            "schema": "public",
            "django_app": "products",
            "description": "Recently viewed products for browsing history.",
            "primary_key": {"field": "id", "type": "BigAutoField"},
            "fields": [
                {"name": "id", "type": "BigAutoField", "pk": True, "nullable": False, "description": "High-volume table uses integer PK"},
                {"name": "user", "type": "FK(User)", "nullable": True, "fk_to": "User", "description": "Logged-in user"},
                {"name": "session_id", "type": "CharField(64)", "nullable": True, "description": "Anonymous session ID"},
                {"name": "product", "type": "FK(Product)", "nullable": False, "fk_to": "Product", "description": "Product viewed"},
                {"name": "viewed_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ]
        },
        "StockAlert": {
            "schema": "public",
            "django_app": "products",
            "description": "Back-in-stock notification subscription.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "description": "Subscriber"},
                {"name": "product", "type": "FK(Product)", "nullable": False, "fk_to": "Product", "description": "Product to watch"},
                {"name": "listing", "type": "FK(ProductListing)", "nullable": True, "fk_to": "ProductListing", "description": "Specific listing (null = any)"},
                {"name": "is_active", "type": "BooleanField", "nullable": False, "default": True, "description": "Alert active flag"},
                {"name": "notified_at", "type": "DateTimeField", "nullable": True, "description": "When notification was sent"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ],
            "unique_constraints": [["user", "product", "listing"]]
        },
        "CategoryPreferenceSchema": {
            "schema": "public",
            "django_app": "products",
            "description": "JSON schema defining the recommendation questionnaire for a product category.",
            "primary_key": {"field": "id", "type": "int"},
            "fields": [
                {"name": "id", "type": "IntegerField", "pk": True, "nullable": False, "description": ""},
                {"name": "category", "type": "OneToOneField(Category)", "nullable": False, "fk_to": "Category", "description": "Category this schema belongs to"},
                {"name": "schema", "type": "JSONField", "nullable": False, "description": "JSON Schema for questionnaire fields"},
                {"name": "version", "type": "IntegerField", "nullable": False, "default": 1, "description": "Schema version"},
                {"name": "is_active", "type": "BooleanField", "nullable": False, "default": True, "description": "Whether this schema is in use"}
            ]
        },
        "PriceSnapshot": {
            "schema": "public",
            "django_app": "pricing",
            "description": "TimescaleDB hypertable. Time-series price data — one row per scrape cycle per listing.",
            "primary_key": {"field": ["time", "listing_id"], "type": "composite"},
            "hypertable": True,
            "managed": False,
            "partition_key": "time",
            "fields": [
                {"name": "time", "type": "TimestampTZ", "pk": True, "nullable": False, "description": "Scrape timestamp — hypertable partition key"},
                {"name": "listing", "type": "FK(ProductListing)", "pk": True, "nullable": False, "fk_to": "ProductListing", "description": "Which listing was scraped"},
                {"name": "product", "type": "FK(Product)", "nullable": False, "fk_to": "Product", "description": "Canonical product (denormalized)"},
                {"name": "marketplace", "type": "FK(Marketplace)", "nullable": False, "fk_to": "Marketplace", "description": "Source marketplace (denormalized)"},
                {"name": "price", "type": "DecimalField(12,2)", "nullable": False, "description": "Selling price (paisa)"},
                {"name": "mrp", "type": "DecimalField(12,2)", "nullable": True, "description": "MRP (paisa)"},
                {"name": "discount_pct", "type": "DecimalField(5,2)", "nullable": True, "description": "Computed discount percentage"},
                {"name": "in_stock", "type": "BooleanField", "nullable": False, "description": "Stock status at scrape time"},
                {"name": "seller_name", "type": "CharField(300)", "nullable": True, "description": "Seller name at scrape time"},
                {"name": "source", "type": "CharField(20)", "nullable": True, "description": "Data source (scraper, api, manual)"}
            ]
        },
        "MarketplaceOffer": {
            "schema": "public",
            "django_app": "pricing",
            "description": "Bank/card/wallet offers scraped from marketplace pages for smart payment optimization.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "marketplace", "type": "FK(Marketplace)", "nullable": False, "fk_to": "Marketplace", "description": "Source marketplace"},
                {"name": "scope_type", "type": "CharField(20)", "nullable": False, "description": "Offer scope: product, category, marketplace"},
                {"name": "product", "type": "FK(Product)", "nullable": True, "fk_to": "Product", "description": "Product-specific offer target"},
                {"name": "listing", "type": "FK(ProductListing)", "nullable": True, "fk_to": "ProductListing", "description": "Listing-specific offer target"},
                {"name": "category", "type": "FK(Category)", "nullable": True, "fk_to": "Category", "description": "Category-wide offer target"},
                {"name": "offer_type", "type": "CharField(20)", "nullable": False, "description": "Offer mechanism (bank_offer, coupon, etc.)"},
                {"name": "title", "type": "CharField(300)", "nullable": False, "description": "Offer headline"},
                {"name": "description", "type": "TextField", "nullable": True, "description": "Full offer terms"},
                {"name": "bank_slug", "type": "CharField(50)", "nullable": True, "description": "Bank identifier for matching"},
                {"name": "card_type", "type": "CharField(10)", "nullable": True, "description": "credit or debit"},
                {"name": "card_network", "type": "CharField(20)", "nullable": True, "description": "Visa / Mastercard / RuPay"},
                {"name": "card_variants", "type": "JSONField", "nullable": True, "description": "Specific card variants eligible"},
                {"name": "wallet_provider", "type": "CharField(50)", "nullable": True, "description": "E-wallet name"},
                {"name": "membership_type", "type": "CharField(50)", "nullable": True, "description": "Membership requirement"},
                {"name": "coupon_code", "type": "CharField(50)", "nullable": True, "description": "Coupon code"},
                {"name": "discount_type", "type": "CharField(15)", "nullable": False, "choices": ["flat","percent","cashback","no_cost_emi"], "description": "Discount mechanism"},
                {"name": "discount_value", "type": "DecimalField(10,2)", "nullable": False, "description": "Discount amount"},
                {"name": "max_discount", "type": "DecimalField(10,2)", "nullable": True, "description": "Maximum discount cap (paisa)"},
                {"name": "min_purchase", "type": "DecimalField(12,2)", "nullable": True, "description": "Minimum purchase amount (paisa)"},
                {"name": "emi_tenures", "type": "JSONField", "nullable": True, "description": "Available EMI tenure options in months"},
                {"name": "emi_interest_rate", "type": "DecimalField(5,2)", "nullable": True, "description": "EMI interest rate"},
                {"name": "emi_processing_fee", "type": "DecimalField(10,2)", "nullable": True, "description": "EMI processing fee (paisa)"},
                {"name": "valid_from", "type": "DateTimeField", "nullable": True, "description": "Offer start date"},
                {"name": "valid_until", "type": "DateTimeField", "nullable": True, "description": "Offer end date"},
                {"name": "is_active", "type": "BooleanField", "nullable": False, "default": True, "description": "Current validity"},
                {"name": "stackable", "type": "BooleanField", "nullable": False, "default": False, "description": "Can combine with other offers"},
                {"name": "source", "type": "CharField(50)", "nullable": True, "description": "Where offer was scraped from"},
                {"name": "last_verified_at", "type": "DateTimeField", "nullable": True, "description": "Last verification timestamp"},
                {"name": "terms_conditions", "type": "TextField", "nullable": True, "description": "Full T&C text"}
            ]
        },
        "PriceAlert": {
            "schema": "public",
            "django_app": "pricing",
            "description": "User-configured price drop notifications.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "description": "Alert owner"},
                {"name": "product", "type": "FK(Product)", "nullable": False, "fk_to": "Product", "description": "Target product"},
                {"name": "target_price", "type": "DecimalField(12,2)", "nullable": False, "description": "Price threshold (paisa)"},
                {"name": "current_price", "type": "DecimalField(12,2)", "nullable": True, "description": "Current best price at creation (paisa)"},
                {"name": "marketplace", "type": "FK(Marketplace)", "nullable": True, "fk_to": "Marketplace", "description": "Specific marketplace (null = any)"},
                {"name": "is_active", "type": "BooleanField", "nullable": False, "default": True, "description": "Alert active flag"},
                {"name": "is_triggered", "type": "BooleanField", "nullable": False, "default": False, "description": "Whether alert has fired"},
                {"name": "triggered_at", "type": "DateTimeField", "nullable": True, "description": "When price dropped below target"},
                {"name": "triggered_price", "type": "DecimalField(12,2)", "nullable": True, "description": "Actual triggered price (paisa)"},
                {"name": "triggered_marketplace", "type": "CharField(30)", "nullable": True, "description": "Marketplace where price dropped"},
                {"name": "notification_sent", "type": "BooleanField", "nullable": False, "default": False, "description": "Whether notification was dispatched"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ],
            "unique_constraints": [["user", "product", "marketplace"]]
        },
        "ClickEvent": {
            "schema": "public",
            "django_app": "pricing",
            "description": "Affiliate click tracking for attribution and analytics.",
            "primary_key": {"field": "id", "type": "BigAutoField"},
            "fields": [
                {"name": "id", "type": "BigAutoField", "pk": True, "nullable": False, "description": "High-volume, integer PK"},
                {"name": "user", "type": "FK(User)", "nullable": True, "fk_to": "User", "description": "Logged-in clicker"},
                {"name": "session_id", "type": "CharField(64)", "nullable": True, "description": "Anonymous session"},
                {"name": "product", "type": "FK(Product)", "nullable": False, "fk_to": "Product", "description": "Product clicked"},
                {"name": "listing", "type": "FK(ProductListing)", "nullable": False, "fk_to": "ProductListing", "description": "Specific listing clicked"},
                {"name": "marketplace", "type": "FK(Marketplace)", "nullable": False, "fk_to": "Marketplace", "description": "Target marketplace"},
                {"name": "source_page", "type": "CharField(50)", "nullable": False, "description": "Page where click happened"},
                {"name": "source_section", "type": "CharField(50)", "nullable": True, "description": "UI section"},
                {"name": "affiliate_url", "type": "URLField(2000)", "nullable": False, "description": "Full affiliate URL"},
                {"name": "affiliate_tag", "type": "CharField(50)", "nullable": False, "description": "Affiliate tag used"},
                {"name": "sub_tag", "type": "CharField(100)", "nullable": True, "description": "Sub-tracking tag"},
                {"name": "purchase_confirmed", "type": "BooleanField", "nullable": False, "default": False, "description": "Purchase confirmed via email intel"},
                {"name": "confirmation_source", "type": "CharField(30)", "nullable": True, "description": "How purchase was confirmed"},
                {"name": "confirmed_at", "type": "DateTimeField", "nullable": True, "description": "Purchase confirmation time"},
                {"name": "price_at_click", "type": "DecimalField(12,2)", "nullable": True, "description": "Price at click time (paisa)"},
                {"name": "device_type", "type": "CharField(20)", "nullable": True, "description": "mobile, tablet, desktop"},
                {"name": "referrer", "type": "URLField", "nullable": True, "description": "HTTP referrer"},
                {"name": "ip_hash", "type": "CharField(64)", "nullable": True, "description": "SHA-256 hashed IP (privacy-safe)"},
                {"name": "user_agent_hash", "type": "CharField(64)", "nullable": True, "description": "SHA-256 hashed UA string"},
                {"name": "clicked_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ]
        },
        "BackfillProduct": {
            "schema": "public",
            "django_app": "pricing",
            "description": "Staging table for multi-phase price history backfill.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "product_listing", "type": "FK(ProductListing)", "nullable": True, "fk_to": "ProductListing", "description": "Linked listing (after phase 1 match)"},
                {"name": "ph_code", "type": "CharField(50)", "nullable": True, "description": "PriceHistory.co product code"},
                {"name": "ph_url", "type": "URLField(2000)", "nullable": True, "description": "PriceHistory.co URL"},
                {"name": "marketplace_slug", "type": "CharField(30)", "nullable": False, "description": "Source marketplace"},
                {"name": "external_id", "type": "CharField(100)", "nullable": False, "description": "External product ID"},
                {"name": "marketplace_url", "type": "URLField(2000)", "nullable": True, "description": "Original marketplace URL"},
                {"name": "title", "type": "CharField(500)", "nullable": False, "description": "Product title"},
                {"name": "brand_name", "type": "CharField(200)", "nullable": True, "description": "Brand name"},
                {"name": "image_url", "type": "URLField(2000)", "nullable": True, "description": "Product image"},
                {"name": "price_data_points", "type": "IntegerField", "nullable": False, "default": 0, "description": "Number of historical price points"},
                {"name": "history_from", "type": "DateField", "nullable": True, "description": "Earliest price data point"},
                {"name": "history_to", "type": "DateField", "nullable": True, "description": "Latest price data point"},
                {"name": "bh_prediction_days", "type": "IntegerField", "nullable": True, "description": "BuyHatke: days to next drop"},
                {"name": "bh_prediction_weeks", "type": "IntegerField", "nullable": True, "description": "BuyHatke: weeks"},
                {"name": "bh_prediction_months", "type": "IntegerField", "nullable": True, "description": "BuyHatke: months"},
                {"name": "bh_popularity", "type": "DecimalField(5,2)", "nullable": True, "description": "BuyHatke popularity index"},
                {"name": "min_price", "type": "DecimalField(12,2)", "nullable": True, "description": "Historical minimum (paisa)"},
                {"name": "max_price", "type": "DecimalField(12,2)", "nullable": True, "description": "Historical maximum (paisa)"},
                {"name": "min_price_date", "type": "DateField", "nullable": True, "description": "Date of minimum"},
                {"name": "max_price_date", "type": "DateField", "nullable": True, "description": "Date of maximum"},
                {"name": "status", "type": "CharField(20)", "nullable": False, "default": "discovered", "choices": ["discovered","bh_filled","ph_extended","done","failed","skipped"], "description": "Pipeline status"},
                {"name": "error_message", "type": "TextField", "nullable": True, "description": "Last error details"},
                {"name": "retry_count", "type": "IntegerField", "nullable": False, "default": 0, "description": "Processing retries"}
            ]
        },
        "Review": {
            "schema": "public",
            "django_app": "reviews",
            "description": "Product reviews — scraped from marketplaces and written natively on Whydud.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "listing", "type": "FK(ProductListing)", "nullable": True, "fk_to": "ProductListing", "description": "Marketplace listing (null for Whydud reviews)"},
                {"name": "product", "type": "FK(Product)", "nullable": False, "fk_to": "Product", "description": "Canonical product"},
                {"name": "user", "type": "FK(User)", "nullable": True, "fk_to": "User", "description": "Whydud user (null for scraped)"},
                {"name": "source", "type": "CharField(10)", "nullable": False, "choices": ["scraped","whydud"], "description": "Review origin"},
                {"name": "external_review_id", "type": "CharField(200)", "nullable": True, "description": "Marketplace's review ID"},
                {"name": "reviewer_name", "type": "CharField(200)", "nullable": True, "description": "Display name"},
                {"name": "marketplace", "type": "FK(Marketplace)", "nullable": True, "fk_to": "Marketplace", "description": "Source marketplace"},
                {"name": "rating", "type": "IntegerField", "nullable": False, "description": "Star rating (1-5)"},
                {"name": "title", "type": "CharField(500)", "nullable": True, "description": "Review headline"},
                {"name": "body", "type": "TextField", "nullable": True, "description": "Full review text"},
                {"name": "is_verified_purchase", "type": "BooleanField", "nullable": False, "default": False, "description": "Marketplace verified purchase badge"},
                {"name": "sentiment_score", "type": "DecimalField(3,2)", "nullable": True, "description": "AI sentiment (-1.00 to 1.00)"},
                {"name": "sentiment_label", "type": "CharField(20)", "nullable": True, "description": "positive, negative, neutral, mixed"},
                {"name": "credibility_score", "type": "DecimalField(5,2)", "nullable": True, "description": "Fraud detection score (0.00-100.00)"},
                {"name": "fraud_flags", "type": "JSONField", "nullable": False, "default": [], "description": "Detected fraud indicators"},
                {"name": "is_flagged", "type": "BooleanField", "nullable": False, "default": False, "description": "Flagged for moderation"},
                {"name": "is_published", "type": "BooleanField", "nullable": False, "default": True, "description": "Visibility flag"},
                {"name": "upvotes", "type": "IntegerField", "nullable": False, "default": 0, "description": "Community upvotes"},
                {"name": "downvotes", "type": "IntegerField", "nullable": False, "default": 0, "description": "Community downvotes"},
                {"name": "vote_score", "type": "IntegerField", "nullable": False, "default": 0, "description": "Net score (upvotes - downvotes)"},
                {"name": "review_date", "type": "DateField", "nullable": True, "description": "When review was written"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ]
        },
        "ReviewVote": {
            "schema": "public", "django_app": "reviews",
            "description": "Community voting on reviews.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "review", "type": "FK(Review)", "nullable": False, "fk_to": "Review", "on_delete": "CASCADE", "description": "Voted review"},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "on_delete": "CASCADE", "description": "Voter"},
                {"name": "vote", "type": "IntegerField", "nullable": False, "description": "1 = upvote, -1 = downvote"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ],
            "unique_constraints": [["review", "user"]]
        },
        "ReviewerProfile": {
            "schema": "public", "django_app": "reviews",
            "description": "Gamification layer for review authors.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "OneToOneField(User)", "nullable": False, "fk_to": "User", "description": "Profile owner"},
                {"name": "total_reviews", "type": "IntegerField", "nullable": False, "default": 0, "description": "Lifetime reviews"},
                {"name": "total_upvotes_received", "type": "IntegerField", "nullable": False, "default": 0, "description": "Lifetime upvotes"},
                {"name": "reviewer_level", "type": "CharField(10)", "nullable": False, "default": "bronze", "choices": ["bronze","silver","gold","platinum"], "description": "Gamification level"},
                {"name": "is_top_reviewer", "type": "BooleanField", "nullable": False, "default": False, "description": "Top reviewer badge"}
            ]
        },
        "DudScoreConfig": {
            "schema": "scoring", "django_app": "scoring",
            "description": "Versioned weight configuration for DudScore algorithm. Only one active config at a time.",
            "primary_key": {"field": "id", "type": "int"},
            "fields": [
                {"name": "id", "type": "IntegerField", "pk": True, "nullable": False, "description": ""},
                {"name": "version", "type": "IntegerField", "unique": True, "nullable": False, "description": "Configuration version number"},
                {"name": "w_sentiment", "type": "DecimalField(3,2)", "nullable": False, "description": "Weight: review sentiment"},
                {"name": "w_rating_quality", "type": "DecimalField(3,2)", "nullable": False, "description": "Weight: rating quality/distribution"},
                {"name": "w_price_value", "type": "DecimalField(3,2)", "nullable": False, "description": "Weight: price-to-value ratio"},
                {"name": "w_review_credibility", "type": "DecimalField(3,2)", "nullable": False, "description": "Weight: review authenticity"},
                {"name": "w_price_stability", "type": "DecimalField(3,2)", "nullable": False, "description": "Weight: price stability"},
                {"name": "w_return_signal", "type": "DecimalField(3,2)", "nullable": False, "description": "Weight: return/refund signals"},
                {"name": "is_active", "type": "BooleanField", "nullable": False, "default": False, "description": "Whether this is the live config"},
                {"name": "created_by", "type": "FK(User)", "nullable": True, "fk_to": "User", "description": "Admin who created config"}
            ]
        },
        "DudScoreHistory": {
            "schema": "scoring", "django_app": "scoring",
            "description": "TimescaleDB hypertable for DudScore evolution over time.",
            "primary_key": {"field": ["time", "product_id"], "type": "composite"},
            "hypertable": True, "managed": False, "partition_key": "time",
            "fields": [
                {"name": "time", "type": "TimestampTZ", "pk": True, "nullable": False, "description": "Computation timestamp"},
                {"name": "product", "type": "FK(Product)", "pk": True, "nullable": False, "fk_to": "Product", "description": "Scored product"},
                {"name": "score", "type": "DecimalField(5,2)", "nullable": False, "description": "DudScore (0.00-100.00)"},
                {"name": "config_version", "type": "IntegerField", "nullable": False, "description": "DudScoreConfig version used"},
                {"name": "component_scores", "type": "JSONField", "nullable": False, "default": {}, "description": "Score breakdown per component"}
            ]
        },
        "BrandTrustScore": {
            "schema": "scoring", "django_app": "scoring",
            "description": "Aggregated brand-level trust metric. Recomputed weekly.",
            "primary_key": {"field": "id", "type": "int"},
            "fields": [
                {"name": "id", "type": "IntegerField", "pk": True, "nullable": False, "description": ""},
                {"name": "brand", "type": "OneToOneField(Brand)", "nullable": False, "fk_to": "Brand", "description": "Scored brand"},
                {"name": "avg_dud_score", "type": "DecimalField(5,2)", "nullable": False, "description": "Mean DudScore across products"},
                {"name": "trust_tier", "type": "CharField(15)", "nullable": False, "choices": ["excellent","good","average","poor","avoid"], "description": "Tier label"},
                {"name": "computed_at", "type": "DateTimeField", "nullable": False, "description": "Last computation time"}
            ]
        },
        "InboxEmail": {
            "schema": "email_intel", "django_app": "email_intel",
            "description": "User's shopping emails. Body content encrypted at rest with AES-256-GCM.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "description": "Email owner"},
                {"name": "whydud_email", "type": "FK(WhydudEmail)", "nullable": True, "fk_to": "WhydudEmail", "description": "Destination shopping email"},
                {"name": "message_id", "type": "CharField(255)", "unique": True, "nullable": False, "description": "RFC 2822 Message-ID"},
                {"name": "sender_address", "type": "EmailField", "nullable": False, "description": "From address"},
                {"name": "subject", "type": "CharField(500)", "nullable": False, "description": "Email subject"},
                {"name": "body_text_encrypted", "type": "BinaryField", "nullable": True, "encrypted": "AES-256-GCM", "description": "Encrypted plain text body"},
                {"name": "body_html_encrypted", "type": "BinaryField", "nullable": True, "encrypted": "AES-256-GCM", "description": "Encrypted HTML body"},
                {"name": "category", "type": "CharField(20)", "nullable": False, "default": "other", "choices": ["order","shipping","delivery","refund","return","subscription","promo","otp","other"], "description": "AI-classified category"},
                {"name": "parse_status", "type": "CharField(20)", "nullable": False, "default": "pending", "choices": ["pending","parsed","failed","failed_permanent","skipped"], "description": "Parsing pipeline status"},
                {"name": "received_at", "type": "DateTimeField", "nullable": False, "description": "When email was received"}
            ]
        },
        "EmailSource": {
            "schema": "email_intel", "django_app": "email_intel",
            "description": "All email sources connected by a user.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "description": ""},
                {"name": "source_type", "type": "CharField(15)", "nullable": False, "choices": ["whydud","gmail","outlook","forwarding"], "description": "Source type"},
                {"name": "email_address", "type": "EmailField", "nullable": False, "description": "Connected email"},
                {"name": "sync_status", "type": "CharField(15)", "nullable": False, "default": "active", "choices": ["active","paused","error","disconnected"], "description": "Sync health"}
            ],
            "unique_constraints": [["user", "email_address"]]
        },
        "ParsedOrder": {
            "schema": "email_intel", "django_app": "email_intel",
            "description": "Purchase orders extracted from emails.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "description": ""},
                {"name": "source", "type": "FK(EmailSource)", "nullable": False, "fk_to": "EmailSource", "description": "Parsed from"},
                {"name": "order_id", "type": "CharField(100)", "nullable": False, "description": "Marketplace order ID"},
                {"name": "marketplace", "type": "FK(Marketplace)", "nullable": True, "fk_to": "Marketplace", "description": "Detected marketplace"},
                {"name": "product_name", "type": "CharField(500)", "nullable": False, "description": "Product name from email"},
                {"name": "total_amount", "type": "DecimalField(12,2)", "nullable": False, "description": "Total order amount (paisa)"},
                {"name": "matched_product", "type": "FK(Product)", "nullable": True, "fk_to": "Product", "description": "Matched canonical product"},
                {"name": "order_date", "type": "DateField", "nullable": False, "description": "Order placement date"}
            ],
            "unique_constraints": [["user", "email_message_id"]]
        },
        "RefundTracking": {
            "schema": "email_intel", "django_app": "email_intel",
            "description": "Tracks refund processing.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "description": ""},
                {"name": "order", "type": "FK(ParsedOrder)", "nullable": False, "fk_to": "ParsedOrder", "description": "Related order"},
                {"name": "status", "type": "CharField(20)", "nullable": False, "choices": ["initiated","processing","completed","failed"], "description": "Refund status"},
                {"name": "refund_amount", "type": "DecimalField(12,2)", "nullable": False, "description": "Amount (paisa)"},
                {"name": "delay_days", "type": "IntegerField", "nullable": False, "default": 0, "description": "Days past expected date"}
            ]
        },
        "ReturnWindow": {
            "schema": "email_intel", "django_app": "email_intel",
            "description": "Tracks return eligibility and sends alerts.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "description": ""},
                {"name": "order", "type": "FK(ParsedOrder)", "nullable": False, "fk_to": "ParsedOrder", "description": "Related order"},
                {"name": "window_end_date", "type": "DateField", "nullable": False, "description": "Return window expiry"},
                {"name": "alert_sent_3day", "type": "BooleanField", "nullable": False, "default": False, "description": "3-day warning sent"},
                {"name": "alert_sent_1day", "type": "BooleanField", "nullable": False, "default": False, "description": "1-day warning sent"}
            ]
        },
        "DetectedSubscription": {
            "schema": "email_intel", "django_app": "email_intel",
            "description": "Recurring subscriptions detected from email patterns.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "description": ""},
                {"name": "service_name", "type": "CharField(200)", "nullable": False, "description": "Service name (e.g. Netflix)"},
                {"name": "amount", "type": "DecimalField(12,2)", "nullable": False, "description": "Amount (paisa)"},
                {"name": "billing_cycle", "type": "CharField(20)", "nullable": False, "description": "monthly, quarterly, yearly"},
                {"name": "is_active", "type": "BooleanField", "nullable": False, "default": True, "description": "Active flag"}
            ]
        },
        "Wishlist": {
            "schema": "users", "django_app": "wishlists",
            "description": "User's product wishlists with optional public sharing.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "on_delete": "CASCADE", "description": "Owner"},
                {"name": "name", "type": "CharField(100)", "nullable": False, "description": "Wishlist name"},
                {"name": "is_default", "type": "BooleanField", "nullable": False, "default": False, "description": "Default wishlist"},
                {"name": "is_public", "type": "BooleanField", "nullable": False, "default": False, "description": "Publicly visible"},
                {"name": "share_slug", "type": "SlugField", "unique": True, "nullable": True, "description": "Public share URL slug"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ]
        },
        "WishlistItem": {
            "schema": "users", "django_app": "wishlists",
            "description": "Product in a wishlist with price tracking.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "wishlist", "type": "FK(Wishlist)", "nullable": False, "fk_to": "Wishlist", "on_delete": "CASCADE", "description": "Parent wishlist"},
                {"name": "product", "type": "FK(Product)", "nullable": False, "fk_to": "Product", "description": "Wishlisted product"},
                {"name": "price_when_added", "type": "DecimalField(12,2)", "nullable": True, "description": "Price at add time (paisa)"},
                {"name": "target_price", "type": "DecimalField(12,2)", "nullable": True, "description": "Desired price for alert (paisa)"},
                {"name": "alert_enabled", "type": "BooleanField", "nullable": False, "default": False, "description": "Price drop alert toggle"},
                {"name": "added_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ],
            "unique_constraints": [["wishlist", "product"]]
        },
        "Deal": {
            "schema": "public", "django_app": "deals",
            "description": "Automatically detected deals based on price analysis.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "product", "type": "FK(Product)", "nullable": False, "fk_to": "Product", "description": "Product on deal"},
                {"name": "listing", "type": "FK(ProductListing)", "nullable": False, "fk_to": "ProductListing", "description": "Listing with deal"},
                {"name": "marketplace", "type": "FK(Marketplace)", "nullable": False, "fk_to": "Marketplace", "description": ""},
                {"name": "deal_type", "type": "CharField(20)", "nullable": False, "choices": ["error_price","lowest_ever","genuine_discount","flash_sale"], "description": "Detection type"},
                {"name": "current_price", "type": "DecimalField(12,2)", "nullable": False, "description": "Deal price (paisa)"},
                {"name": "reference_price", "type": "DecimalField(12,2)", "nullable": False, "description": "Reference price (paisa)"},
                {"name": "discount_pct", "type": "DecimalField(5,2)", "nullable": False, "description": "Discount percentage"},
                {"name": "confidence", "type": "CharField(10)", "nullable": False, "choices": ["high","medium","low"], "description": "Detection confidence"},
                {"name": "is_active", "type": "BooleanField", "nullable": False, "default": True, "description": "Active flag"},
                {"name": "detected_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ]
        },
        "RewardPointsLedger": {
            "schema": "users", "django_app": "rewards",
            "description": "Immutable ledger of all points transactions.",
            "primary_key": {"field": "id", "type": "int"},
            "fields": [
                {"name": "id", "type": "IntegerField", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "description": ""},
                {"name": "points", "type": "IntegerField", "nullable": False, "description": "Positive = earned, Negative = spent"},
                {"name": "action_type", "type": "CharField(50)", "nullable": False, "description": "Action trigger (review_written, referral, etc.)"},
                {"name": "description", "type": "CharField(200)", "nullable": False, "description": "Human-readable description"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ]
        },
        "RewardBalance": {
            "schema": "users", "django_app": "rewards",
            "description": "Denormalized current balance (computed from ledger).",
            "primary_key": {"field": "user_id", "type": "UUID (OneToOne)"},
            "fields": [
                {"name": "user", "type": "OneToOneField(User)", "pk": True, "nullable": False, "fk_to": "User", "description": "Balance owner"},
                {"name": "total_earned", "type": "IntegerField", "nullable": False, "default": 0, "description": "Lifetime earned"},
                {"name": "total_spent", "type": "IntegerField", "nullable": False, "default": 0, "description": "Lifetime spent"},
                {"name": "total_expired", "type": "IntegerField", "nullable": False, "default": 0, "description": "Lifetime expired"},
                {"name": "current_balance", "type": "IntegerField", "nullable": False, "default": 0, "description": "Available points"},
                {"name": "updated_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ]
        },
        "GiftCardCatalog": {
            "schema": "users", "django_app": "rewards",
            "description": "Available gift card brands for redemption.",
            "primary_key": {"field": "id", "type": "int"},
            "fields": [
                {"name": "id", "type": "IntegerField", "pk": True, "nullable": False, "description": ""},
                {"name": "brand_name", "type": "CharField(100)", "nullable": False, "description": "Gift card brand"},
                {"name": "brand_slug", "type": "SlugField", "unique": True, "nullable": False, "description": "URL-safe identifier"},
                {"name": "denominations", "type": "JSONField", "nullable": False, "default": [], "description": "Available amounts"},
                {"name": "is_active", "type": "BooleanField", "nullable": False, "default": True, "description": "Available for redemption"}
            ]
        },
        "GiftCardRedemption": {
            "schema": "users", "django_app": "rewards",
            "description": "Gift card redemption transactions.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "description": ""},
                {"name": "catalog", "type": "FK(GiftCardCatalog)", "nullable": False, "fk_to": "GiftCardCatalog", "description": ""},
                {"name": "points_spent", "type": "IntegerField", "nullable": False, "description": "Points deducted"},
                {"name": "status", "type": "CharField(15)", "nullable": False, "choices": ["pending","fulfilled","failed","cancelled"], "description": "Fulfillment status"},
                {"name": "gift_card_code", "type": "BinaryField", "nullable": True, "encrypted": "AES-256-GCM", "description": "Encrypted gift card code"}
            ]
        },
        "DiscussionThread": {
            "schema": "community", "django_app": "discussions",
            "description": "Product discussion threads.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "product", "type": "FK(Product)", "nullable": False, "fk_to": "Product", "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "description": "Author"},
                {"name": "thread_type", "type": "CharField(15)", "nullable": False, "choices": ["question","experience","comparison","tip","alert"], "description": "Thread category"},
                {"name": "title", "type": "CharField(300)", "nullable": False, "description": "Thread title"},
                {"name": "body", "type": "TextField", "nullable": False, "description": "Thread body"},
                {"name": "reply_count", "type": "IntegerField", "nullable": False, "default": 0, "description": "Denormalized reply count"},
                {"name": "is_pinned", "type": "BooleanField", "nullable": False, "default": False, "description": "Moderator pinned"},
                {"name": "is_locked", "type": "BooleanField", "nullable": False, "default": False, "description": "No new replies"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ]
        },
        "DiscussionReply": {
            "schema": "community", "django_app": "discussions",
            "description": "Replies to discussion threads. Supports nesting.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "thread", "type": "FK(DiscussionThread)", "nullable": False, "fk_to": "DiscussionThread", "on_delete": "CASCADE", "description": "Parent thread"},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "description": ""},
                {"name": "parent_reply", "type": "FK(DiscussionReply)", "nullable": True, "fk_to": "DiscussionReply", "description": "Parent reply (self-ref for nesting)"},
                {"name": "body", "type": "TextField", "nullable": False, "description": "Reply content"},
                {"name": "is_accepted", "type": "BooleanField", "nullable": False, "default": False, "description": "Accepted answer"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ]
        },
        "DiscussionVote": {
            "schema": "community", "django_app": "discussions",
            "description": "Polymorphic voting for threads and replies.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "FK(User)", "nullable": False, "fk_to": "User", "description": ""},
                {"name": "target_type", "type": "CharField(20)", "nullable": False, "description": "thread or reply"},
                {"name": "target_id", "type": "CharField(50)", "nullable": False, "description": "UUID of voted entity"},
                {"name": "vote", "type": "IntegerField", "nullable": False, "description": "1 = upvote, -1 = downvote"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ],
            "unique_constraints": [["user", "target_type", "target_id"]]
        },
        "TCOModel": {
            "schema": "tco", "django_app": "tco",
            "description": "Defines TCO calculation model for a product category.",
            "primary_key": {"field": "id", "type": "int"},
            "fields": [
                {"name": "id", "type": "IntegerField", "pk": True, "nullable": False, "description": ""},
                {"name": "category", "type": "FK(Category)", "nullable": False, "fk_to": "Category", "description": ""},
                {"name": "name", "type": "CharField(100)", "nullable": False, "description": "Model name"},
                {"name": "version", "type": "IntegerField", "nullable": False, "default": 1, "description": "Model version"},
                {"name": "input_schema", "type": "JSONField", "nullable": False, "default": {}, "description": "JSON Schema for user inputs"},
                {"name": "cost_components", "type": "JSONField", "nullable": False, "default": {}, "description": "Calculation definitions"}
            ],
            "unique_constraints": [["category", "version"]]
        },
        "CityReferenceData": {
            "schema": "tco", "django_app": "tco",
            "description": "Indian city-level reference data for TCO calculations.",
            "primary_key": {"field": "id", "type": "int"},
            "fields": [
                {"name": "id", "type": "IntegerField", "pk": True, "nullable": False, "description": ""},
                {"name": "city_name", "type": "CharField(100)", "nullable": False, "description": "City name"},
                {"name": "state", "type": "CharField(100)", "nullable": False, "description": "Indian state"},
                {"name": "electricity_tariff_residential", "type": "DecimalField(6,2)", "nullable": False, "description": "Rs per kWh residential"},
                {"name": "cooling_days_per_year", "type": "IntegerField", "nullable": False, "default": 0, "description": "Days requiring cooling"},
                {"name": "water_tariff_per_kl", "type": "DecimalField(6,2)", "nullable": True, "description": "Rs per kilolitre"}
            ],
            "unique_constraints": [["city_name", "state"]]
        },
        "UserTCOProfile": {
            "schema": "users", "django_app": "tco",
            "description": "User's saved TCO preferences.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "user", "type": "OneToOneField(User)", "nullable": False, "fk_to": "User", "description": ""},
                {"name": "city", "type": "FK(CityReferenceData)", "nullable": True, "fk_to": "CityReferenceData", "description": "User's city"},
                {"name": "ownership_years", "type": "IntegerField", "nullable": False, "default": 5, "description": "Planned ownership duration"}
            ]
        },
        "SearchLog": {
            "schema": "public", "django_app": "search",
            "description": "Anonymized search analytics (Meilisearch handles actual search).",
            "primary_key": {"field": "id", "type": "BigAutoField"},
            "fields": [
                {"name": "id", "type": "BigAutoField", "pk": True, "nullable": False, "description": ""},
                {"name": "query", "type": "CharField(500)", "nullable": False, "description": "Search query"},
                {"name": "results_count", "type": "IntegerField", "nullable": False, "default": 0, "description": "Results returned"},
                {"name": "latency_ms", "type": "IntegerField", "nullable": False, "default": 0, "description": "Response time (ms)"},
                {"name": "filters_used", "type": "JSONField", "nullable": False, "default": {}, "description": "Applied filters"},
                {"name": "user_id", "type": "UUIDField", "nullable": True, "description": "Searcher (null = anonymous)"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ]
        },
        "ScraperJob": {
            "schema": "public", "django_app": "scraping",
            "description": "Individual scraper execution records.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "marketplace", "type": "FK(Marketplace)", "nullable": False, "fk_to": "Marketplace", "description": ""},
                {"name": "spider_name", "type": "CharField(50)", "nullable": False, "description": "Scrapy spider class"},
                {"name": "status", "type": "CharField(15)", "nullable": False, "default": "queued", "choices": ["queued","running","completed","failed","partial"], "description": "Execution status"},
                {"name": "items_scraped", "type": "IntegerField", "nullable": False, "default": 0, "description": "Items scraped"},
                {"name": "items_failed", "type": "IntegerField", "nullable": False, "default": 0, "description": "Items failed"},
                {"name": "triggered_by", "type": "CharField(15)", "nullable": False, "default": "scheduled", "choices": ["scheduled","adhoc"], "description": "Trigger type"}
            ]
        },
        "AuditLog": {
            "schema": "admin", "django_app": "admin_tools",
            "description": "Immutable record of every admin/moderator action.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "admin_user", "type": "FK(User)", "nullable": False, "fk_to": "User", "description": "Admin who acted"},
                {"name": "action", "type": "CharField(20)", "nullable": False, "choices": ["create","update","delete","approve","reject","suspend","restore","config_change"], "description": "Action type"},
                {"name": "target_type", "type": "CharField(50)", "nullable": False, "description": "Affected model name"},
                {"name": "target_id", "type": "CharField(50)", "nullable": False, "description": "Affected entity ID"},
                {"name": "old_value", "type": "JSONField", "nullable": True, "description": "State before change"},
                {"name": "new_value", "type": "JSONField", "nullable": True, "description": "State after change"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ]
        },
        "ModerationQueue": {
            "schema": "admin", "django_app": "admin_tools",
            "description": "Items pending moderator review.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "item_type", "type": "CharField(20)", "nullable": False, "choices": ["review","discussion","user"], "description": "Item category"},
                {"name": "item_id", "type": "CharField(50)", "nullable": False, "description": "Item UUID"},
                {"name": "reason", "type": "TextField", "nullable": False, "description": "Why flagged"},
                {"name": "status", "type": "CharField(15)", "nullable": False, "default": "pending", "choices": ["pending","approved","rejected"], "description": "Moderation status"},
                {"name": "assigned_to", "type": "FK(User)", "nullable": True, "fk_to": "User", "description": "Assigned moderator"},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ]
        },
        "ScraperRun": {
            "schema": "admin", "django_app": "admin_tools",
            "description": "Aggregated scraper execution records for admin dashboard.",
            "primary_key": {"field": "id", "type": "UUID"},
            "fields": [
                {"name": "id", "type": "UUID", "pk": True, "nullable": False, "description": ""},
                {"name": "marketplace", "type": "FK(Marketplace)", "nullable": False, "fk_to": "Marketplace", "description": ""},
                {"name": "spider_name", "type": "CharField(50)", "nullable": False, "description": ""},
                {"name": "status", "type": "CharField(15)", "nullable": False, "choices": ["running","completed","failed","partial"], "description": ""},
                {"name": "items_scraped", "type": "IntegerField", "nullable": False, "default": 0, "description": ""},
                {"name": "items_created", "type": "IntegerField", "nullable": False, "default": 0, "description": "New records"},
                {"name": "items_updated", "type": "IntegerField", "nullable": False, "default": 0, "description": "Updated records"}
            ]
        },
        "SiteConfig": {
            "schema": "admin", "django_app": "admin_tools",
            "description": "Runtime-tuneable configuration store.",
            "primary_key": {"field": "key", "type": "CharField(100)"},
            "fields": [
                {"name": "key", "type": "CharField(100)", "pk": True, "unique": True, "nullable": False, "description": "Dot-separated config key"},
                {"name": "value", "type": "JSONField", "nullable": False, "description": "Configuration value (any JSON)"},
                {"name": "updated_by", "type": "FK(User)", "nullable": True, "fk_to": "User", "description": "Last modifier"},
                {"name": "updated_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""},
                {"name": "created_at", "type": "DateTimeField", "nullable": False, "auto": True, "description": ""}
            ]
        }
    },
    "relationships": [
        {"from": "User", "to": "WhydudEmail", "type": "1:N", "fk_field": "user_id", "label": "has"},
        {"from": "User", "to": "OAuthConnection", "type": "1:N", "fk_field": "user_id", "label": "connects"},
        {"from": "User", "to": "PaymentMethod", "type": "1:N", "fk_field": "user_id", "label": "owns"},
        {"from": "User", "to": "Notification", "type": "1:N", "fk_field": "user_id", "label": "receives"},
        {"from": "User", "to": "NotificationPreference", "type": "1:N", "fk_field": "user_id", "label": "configures"},
        {"from": "User", "to": "MarketplacePreference", "type": "1:N", "fk_field": "user_id", "label": "sets"},
        {"from": "User", "to": "PurchasePreference", "type": "1:N", "fk_field": "user_id", "label": "answers"},
        {"from": "User", "to": "ReviewerProfile", "type": "1:1", "fk_field": "user_id", "label": "has profile"},
        {"from": "User", "to": "RewardBalance", "type": "1:1", "fk_field": "user_id", "label": "has balance"},
        {"from": "User", "to": "UserTCOProfile", "type": "1:1", "fk_field": "user_id", "label": "has TCO profile"},
        {"from": "Brand", "to": "Product", "type": "1:N", "fk_field": "brand_id", "label": "produces"},
        {"from": "Category", "to": "Product", "type": "1:N", "fk_field": "category_id", "label": "contains"},
        {"from": "Category", "to": "Category", "type": "self-ref", "fk_field": "parent_id", "label": "parent-child"},
        {"from": "Category", "to": "CategoryPreferenceSchema", "type": "1:1", "fk_field": "category_id", "label": "has schema"},
        {"from": "Product", "to": "ProductListing", "type": "1:N", "fk_field": "product_id", "label": "listed as"},
        {"from": "Marketplace", "to": "ProductListing", "type": "1:N", "fk_field": "marketplace_id", "label": "hosts"},
        {"from": "Seller", "to": "ProductListing", "type": "1:N", "fk_field": "seller_id", "label": "sells"},
        {"from": "Marketplace", "to": "Seller", "type": "1:N", "fk_field": "marketplace_id", "label": "has sellers"},
        {"from": "ProductListing", "to": "PriceSnapshot", "type": "1:N", "fk_field": "listing_id", "label": "price tracked"},
        {"from": "Product", "to": "PriceSnapshot", "type": "1:N", "fk_field": "product_id", "label": "price history"},
        {"from": "ProductListing", "to": "MarketplaceOffer", "type": "1:N", "fk_field": "listing_id", "label": "has offers"},
        {"from": "Product", "to": "MarketplaceOffer", "type": "1:N", "fk_field": "product_id", "label": "has offers"},
        {"from": "User", "to": "PriceAlert", "type": "1:N", "fk_field": "user_id", "label": "sets"},
        {"from": "Product", "to": "PriceAlert", "type": "1:N", "fk_field": "product_id", "label": "monitored by"},
        {"from": "User", "to": "ClickEvent", "type": "1:N", "fk_field": "user_id", "label": "clicks"},
        {"from": "ProductListing", "to": "ClickEvent", "type": "1:N", "fk_field": "listing_id", "label": "clicked"},
        {"from": "ProductListing", "to": "BackfillProduct", "type": "1:1", "fk_field": "product_listing_id", "label": "backfill"},
        {"from": "ProductListing", "to": "Review", "type": "1:N", "fk_field": "listing_id", "label": "reviewed on"},
        {"from": "Product", "to": "Review", "type": "1:N", "fk_field": "product_id", "label": "reviewed"},
        {"from": "User", "to": "Review", "type": "1:N", "fk_field": "user_id", "label": "writes"},
        {"from": "Review", "to": "ReviewVote", "type": "1:N", "fk_field": "review_id", "label": "voted on"},
        {"from": "User", "to": "ReviewVote", "type": "1:N", "fk_field": "user_id", "label": "votes"},
        {"from": "Product", "to": "DudScoreHistory", "type": "1:N", "fk_field": "product_id", "label": "scored"},
        {"from": "Brand", "to": "BrandTrustScore", "type": "1:1", "fk_field": "brand_id", "label": "trusted"},
        {"from": "DudScoreConfig", "to": "DudScoreHistory", "type": "1:N", "fk_field": "config_version", "label": "uses config"},
        {"from": "User", "to": "InboxEmail", "type": "1:N", "fk_field": "user_id", "label": "receives"},
        {"from": "WhydudEmail", "to": "InboxEmail", "type": "1:N", "fk_field": "whydud_email_id", "label": "delivered to"},
        {"from": "User", "to": "EmailSource", "type": "1:N", "fk_field": "user_id", "label": "connects"},
        {"from": "OAuthConnection", "to": "EmailSource", "type": "1:N", "fk_field": "oauth_connection_id", "label": "authenticates"},
        {"from": "User", "to": "ParsedOrder", "type": "1:N", "fk_field": "user_id", "label": "ordered"},
        {"from": "EmailSource", "to": "ParsedOrder", "type": "1:N", "fk_field": "source_id", "label": "parsed from"},
        {"from": "ParsedOrder", "to": "Product", "type": "N:1", "fk_field": "matched_product_id", "label": "matched to"},
        {"from": "ParsedOrder", "to": "RefundTracking", "type": "1:1", "fk_field": "order_id", "label": "refunded"},
        {"from": "ParsedOrder", "to": "ReturnWindow", "type": "1:1", "fk_field": "order_id", "label": "return window"},
        {"from": "User", "to": "DetectedSubscription", "type": "1:N", "fk_field": "user_id", "label": "subscribed"},
        {"from": "User", "to": "Wishlist", "type": "1:N", "fk_field": "user_id", "label": "creates"},
        {"from": "Wishlist", "to": "WishlistItem", "type": "1:N", "fk_field": "wishlist_id", "label": "contains"},
        {"from": "Product", "to": "WishlistItem", "type": "1:N", "fk_field": "product_id", "label": "wishlisted"},
        {"from": "Product", "to": "Deal", "type": "1:N", "fk_field": "product_id", "label": "has deals"},
        {"from": "ProductListing", "to": "Deal", "type": "1:N", "fk_field": "listing_id", "label": "deal on"},
        {"from": "Marketplace", "to": "Deal", "type": "1:N", "fk_field": "marketplace_id", "label": "marketplace"},
        {"from": "User", "to": "RewardPointsLedger", "type": "1:N", "fk_field": "user_id", "label": "earns/spends"},
        {"from": "User", "to": "GiftCardRedemption", "type": "1:N", "fk_field": "user_id", "label": "redeems"},
        {"from": "GiftCardCatalog", "to": "GiftCardRedemption", "type": "1:N", "fk_field": "catalog_id", "label": "redeemed from"},
        {"from": "Product", "to": "DiscussionThread", "type": "1:N", "fk_field": "product_id", "label": "discussed"},
        {"from": "User", "to": "DiscussionThread", "type": "1:N", "fk_field": "user_id", "label": "starts"},
        {"from": "DiscussionThread", "to": "DiscussionReply", "type": "1:N", "fk_field": "thread_id", "label": "replied to"},
        {"from": "User", "to": "DiscussionReply", "type": "1:N", "fk_field": "user_id", "label": "replies"},
        {"from": "DiscussionReply", "to": "DiscussionReply", "type": "self-ref", "fk_field": "parent_reply_id", "label": "nested"},
        {"from": "User", "to": "DiscussionVote", "type": "1:N", "fk_field": "user_id", "label": "votes"},
        {"from": "Category", "to": "TCOModel", "type": "1:N", "fk_field": "category_id", "label": "has TCO model"},
        {"from": "CityReferenceData", "to": "UserTCOProfile", "type": "1:N", "fk_field": "city_id", "label": "city ref"},
        {"from": "User", "to": "CompareSession", "type": "1:N", "fk_field": "user_id", "label": "compares"},
        {"from": "User", "to": "RecentlyViewed", "type": "1:N", "fk_field": "user_id", "label": "views"},
        {"from": "Product", "to": "RecentlyViewed", "type": "1:N", "fk_field": "product_id", "label": "viewed"},
        {"from": "User", "to": "StockAlert", "type": "1:N", "fk_field": "user_id", "label": "alerts on"},
        {"from": "Product", "to": "StockAlert", "type": "1:N", "fk_field": "product_id", "label": "stock watched"},
        {"from": "Marketplace", "to": "ScraperJob", "type": "1:N", "fk_field": "marketplace_id", "label": "scraped"},
        {"from": "Marketplace", "to": "ScraperRun", "type": "1:N", "fk_field": "marketplace_id", "label": "run"},
        {"from": "User", "to": "AuditLog", "type": "1:N", "fk_field": "admin_user_id", "label": "audited"},
        {"from": "User", "to": "ModerationQueue", "type": "1:N", "fk_field": "assigned_to", "label": "moderates"}
    ],
    "hypertables": [
        {"table": "PriceSnapshot", "partition_key": "time", "composite_pk": ["time", "listing_id"], "retention": "configurable", "purpose": "Price tracking per listing per scrape cycle"},
        {"table": "DudScoreHistory", "partition_key": "time", "composite_pk": ["time", "product_id"], "retention": "indefinite", "purpose": "DudScore evolution over time"}
    ],
    "encrypted_fields": [
        {"model": "OAuthConnection", "field": "access_token_encrypted", "algorithm": "AES-256-GCM", "purpose": "OAuth access token"},
        {"model": "OAuthConnection", "field": "refresh_token_encrypted", "algorithm": "AES-256-GCM", "purpose": "OAuth refresh token"},
        {"model": "InboxEmail", "field": "body_text_encrypted", "algorithm": "AES-256-GCM", "purpose": "Email plain text body"},
        {"model": "InboxEmail", "field": "body_html_encrypted", "algorithm": "AES-256-GCM", "purpose": "Email HTML body"},
        {"model": "GiftCardRedemption", "field": "gift_card_code", "algorithm": "AES-256-GCM", "purpose": "Redeemed gift card code"}
    ],
    "statistics": {
        "total_models": 50,
        "total_django_apps": 14,
        "total_schemas": 7,
        "total_hypertables": 2,
        "total_encrypted_fields": 5,
        "total_relationships": 68
    }
}

# Write JSON
json_path = r"c:\Users\rames\Downloads\whydud\platform\whydud\docs\ERD.json"
with open(json_path, "w", encoding="utf-8") as f:
    json.dump(erd, f, indent=2, ensure_ascii=False)
print(f"Written: {json_path}")

# =============================================================================
# DATA DICTIONARY EXCEL
# =============================================================================

wb = Workbook()

# Styles
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
section_font = Font(name="Calibri", bold=True, size=12, color="1E293B")
section_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
table_header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
table_header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
normal_font = Font(name="Calibri", size=10)
pk_font = Font(name="Calibri", size=10, bold=True, color="1E293B")
fk_font = Font(name="Calibri", size=10, color="4DB6AC")
encrypted_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0")
)
wrap_align = Alignment(wrap_text=True, vertical="top")

# ---- Sheet 1: Overview ----
ws_overview = wb.active
ws_overview.title = "Overview"
ws_overview.sheet_properties.tabColor = "1E293B"

overview_headers = ["Schema", "Django App", "Table", "Description", "PK Type", "Hypertable", "Row Estimate"]
for col, h in enumerate(overview_headers, 1):
    cell = ws_overview.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

row = 2
for entity_name, entity in sorted(erd["entities"].items(), key=lambda x: (x[1].get("schema",""), x[0])):
    ws_overview.cell(row=row, column=1, value=entity.get("schema","")).font = normal_font
    ws_overview.cell(row=row, column=2, value=entity.get("django_app","")).font = normal_font
    ws_overview.cell(row=row, column=3, value=entity_name).font = Font(name="Calibri", size=10, bold=True)
    ws_overview.cell(row=row, column=4, value=entity.get("description","")).font = normal_font
    ws_overview.cell(row=row, column=4).alignment = wrap_align
    pk = entity.get("primary_key", {})
    pk_str = str(pk.get("type","")) if isinstance(pk.get("field"), str) else f"composite({', '.join(pk.get('field',[]))})"
    ws_overview.cell(row=row, column=5, value=pk_str).font = normal_font
    ws_overview.cell(row=row, column=6, value="Yes" if entity.get("hypertable") else "").font = normal_font
    ws_overview.cell(row=row, column=7, value="").font = normal_font
    for c in range(1, 8):
        ws_overview.cell(row=row, column=c).border = thin_border
    row += 1

ws_overview.column_dimensions["A"].width = 14
ws_overview.column_dimensions["B"].width = 14
ws_overview.column_dimensions["C"].width = 26
ws_overview.column_dimensions["D"].width = 65
ws_overview.column_dimensions["E"].width = 18
ws_overview.column_dimensions["F"].width = 12
ws_overview.column_dimensions["G"].width = 14
ws_overview.auto_filter.ref = f"A1:G{row-1}"

# ---- Sheet 2: Full Data Dictionary ----
ws_dd = wb.create_sheet("Data Dictionary")
ws_dd.sheet_properties.tabColor = "F97316"

dd_headers = ["Schema", "Django App", "Table", "Field", "Type", "PK", "FK To", "Nullable",
              "Unique", "Default", "Choices / Enum Values", "Encrypted", "Description"]
for col, h in enumerate(dd_headers, 1):
    cell = ws_dd.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

row = 2
for entity_name, entity in sorted(erd["entities"].items(), key=lambda x: (x[1].get("schema",""), x[0])):
    for field in entity.get("fields", []):
        ws_dd.cell(row=row, column=1, value=entity.get("schema","")).font = normal_font
        ws_dd.cell(row=row, column=2, value=entity.get("django_app","")).font = normal_font
        ws_dd.cell(row=row, column=3, value=entity_name).font = Font(name="Calibri", size=10, bold=True)
        ws_dd.cell(row=row, column=4, value=field["name"]).font = pk_font if field.get("pk") else (fk_font if field.get("fk_to") else normal_font)
        ws_dd.cell(row=row, column=5, value=field["type"]).font = normal_font
        ws_dd.cell(row=row, column=6, value="PK" if field.get("pk") else "").font = normal_font
        ws_dd.cell(row=row, column=7, value=field.get("fk_to","")).font = fk_font
        ws_dd.cell(row=row, column=8, value="Yes" if field.get("nullable") else "No").font = normal_font
        ws_dd.cell(row=row, column=9, value="Yes" if field.get("unique") else "").font = normal_font
        default_val = field.get("default")
        ws_dd.cell(row=row, column=10, value=str(default_val) if default_val is not None else "").font = normal_font
        choices = field.get("choices", [])
        ws_dd.cell(row=row, column=11, value=", ".join(choices) if choices else "").font = normal_font
        ws_dd.cell(row=row, column=11).alignment = wrap_align
        is_encrypted = field.get("encrypted", "")
        ws_dd.cell(row=row, column=12, value=is_encrypted if is_encrypted else "").font = normal_font
        if is_encrypted:
            ws_dd.cell(row=row, column=12).fill = encrypted_fill
        ws_dd.cell(row=row, column=13, value=field.get("description","")).font = normal_font
        ws_dd.cell(row=row, column=13).alignment = wrap_align
        for c in range(1, 14):
            ws_dd.cell(row=row, column=c).border = thin_border
        row += 1

ws_dd.column_dimensions["A"].width = 13
ws_dd.column_dimensions["B"].width = 13
ws_dd.column_dimensions["C"].width = 24
ws_dd.column_dimensions["D"].width = 28
ws_dd.column_dimensions["E"].width = 22
ws_dd.column_dimensions["F"].width = 5
ws_dd.column_dimensions["G"].width = 22
ws_dd.column_dimensions["H"].width = 9
ws_dd.column_dimensions["I"].width = 8
ws_dd.column_dimensions["J"].width = 14
ws_dd.column_dimensions["K"].width = 40
ws_dd.column_dimensions["L"].width = 14
ws_dd.column_dimensions["M"].width = 55
ws_dd.auto_filter.ref = f"A1:M{row-1}"
ws_dd.freeze_panes = "D2"

# ---- Sheet 3: Relationships ----
ws_rel = wb.create_sheet("Relationships")
ws_rel.sheet_properties.tabColor = "4DB6AC"

rel_headers = ["From Entity", "To Entity", "Relationship Type", "FK Field", "Label"]
for col, h in enumerate(rel_headers, 1):
    cell = ws_rel.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill

row = 2
for rel in erd["relationships"]:
    ws_rel.cell(row=row, column=1, value=rel["from"]).font = normal_font
    ws_rel.cell(row=row, column=2, value=rel["to"]).font = normal_font
    ws_rel.cell(row=row, column=3, value=rel["type"]).font = normal_font
    ws_rel.cell(row=row, column=4, value=rel["fk_field"]).font = normal_font
    ws_rel.cell(row=row, column=5, value=rel["label"]).font = normal_font
    for c in range(1, 6):
        ws_rel.cell(row=row, column=c).border = thin_border
    row += 1

ws_rel.column_dimensions["A"].width = 24
ws_rel.column_dimensions["B"].width = 24
ws_rel.column_dimensions["C"].width = 18
ws_rel.column_dimensions["D"].width = 24
ws_rel.column_dimensions["E"].width = 20
ws_rel.auto_filter.ref = f"A1:E{row-1}"

# ---- Sheet 4: Encrypted Fields ----
ws_enc = wb.create_sheet("Encrypted Fields")
ws_enc.sheet_properties.tabColor = "DC2626"

enc_headers = ["Model", "Field", "Algorithm", "Purpose"]
for col, h in enumerate(enc_headers, 1):
    cell = ws_enc.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")

row = 2
for ef in erd["encrypted_fields"]:
    ws_enc.cell(row=row, column=1, value=ef["model"]).font = normal_font
    ws_enc.cell(row=row, column=2, value=ef["field"]).font = normal_font
    ws_enc.cell(row=row, column=3, value=ef["algorithm"]).font = normal_font
    ws_enc.cell(row=row, column=4, value=ef["purpose"]).font = normal_font
    for c in range(1, 5):
        ws_enc.cell(row=row, column=c).border = thin_border
    row += 1

ws_enc.column_dimensions["A"].width = 22
ws_enc.column_dimensions["B"].width = 28
ws_enc.column_dimensions["C"].width = 16
ws_enc.column_dimensions["D"].width = 30

# ---- Sheet 5: Hypertables ----
ws_ht = wb.create_sheet("Hypertables")
ws_ht.sheet_properties.tabColor = "16A34A"

ht_headers = ["Table", "Partition Key", "Composite PK", "Retention", "Purpose"]
for col, h in enumerate(ht_headers, 1):
    cell = ws_ht.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")

row = 2
for ht in erd["hypertables"]:
    ws_ht.cell(row=row, column=1, value=ht["table"]).font = normal_font
    ws_ht.cell(row=row, column=2, value=ht["partition_key"]).font = normal_font
    ws_ht.cell(row=row, column=3, value=", ".join(ht["composite_pk"])).font = normal_font
    ws_ht.cell(row=row, column=4, value=ht["retention"]).font = normal_font
    ws_ht.cell(row=row, column=5, value=ht["purpose"]).font = normal_font
    for c in range(1, 6):
        ws_ht.cell(row=row, column=c).border = thin_border
    row += 1

ws_ht.column_dimensions["A"].width = 22
ws_ht.column_dimensions["B"].width = 16
ws_ht.column_dimensions["C"].width = 28
ws_ht.column_dimensions["D"].width = 16
ws_ht.column_dimensions["E"].width = 45

# Save Excel
excel_path = r"c:\Users\rames\Downloads\whydud\platform\whydud\docs\DATA-DICTIONARY.xlsx"
wb.save(excel_path)
print(f"Written: {excel_path}")
print("Done!")
